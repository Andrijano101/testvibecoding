"""
data.gov.rs Open Data Scraper (udata API)

Key behavior:
- No hardcoded timestamped file URLs.
- Fetch dataset metadata, choose the newest resource (prefer xlsx, then csv).
- Cache downloads; allow force_refresh.

Outputs:
- data/raw/opendata/parties.json
- data/raw/opendata/parties_source.json
"""

import os
import io
import json
import time
import re
import unicodedata
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import Optional, Dict, List, Tuple

import httpx
import structlog

logger = structlog.get_logger()

UDATA_API = os.getenv("UDATA_API", "https://data.gov.rs/api/1").rstrip("/")
DATA_DIR = os.getenv("DATA_DIR", "./data")
SCRAPE_DELAY = float(os.getenv("SCRAPE_DELAY", "2"))
SCRAPER_TIMEOUT = float(os.getenv("SCRAPER_TIMEOUT", "120"))
USER_AGENT = os.getenv("SCRAPER_UA", "SrpskaTransparentnost/1.0 (+local)")

# Dataset slug for political parties (udata)
PARTIES_DATASET_SLUG = os.getenv("PARTIES_DATASET_SLUG", "politichke-stranke")


@dataclass
class PoliticalPartyRecord:
    party_id: str
    name: str
    name_normalized: str
    abbreviation: Optional[str]
    founded: Optional[str]
    address: Optional[str]
    city: Optional[str]
    leader: Optional[str]
    dissolved: Optional[str]
    source: str = "data.gov.rs"
    source_url: str = "https://data.gov.rs/sr/datasets/politichke-stranke/"
    scraped_at: str = ""

    def __post_init__(self):
        if not self.scraped_at:
            self.scraped_at = datetime.utcnow().isoformat()


def _normalize(text: str) -> str:
    if not text:
        return ""
    # Basic Cyrillic to Latin mapping for Serbian
    cyr = {
        'а':'a','б':'b','в':'v','г':'g','д':'d','ђ':'dj','е':'e','ж':'z','з':'z','и':'i','ј':'j',
        'к':'k','л':'l','љ':'lj','м':'m','н':'n','њ':'nj','о':'o','п':'p','р':'r','с':'s','т':'t',
        'ћ':'c','у':'u','ф':'f','х':'h','ц':'c','ч':'c','џ':'dz','ш':'s',
        'А':'a','Б':'b','В':'v','Г':'g','Д':'d','Ђ':'dj','Е':'e','Ж':'z','З':'z','И':'i','Ј':'j',
        'К':'k','Л':'l','Љ':'lj','М':'m','Н':'n','Њ':'nj','О':'o','П':'p','Р':'r','С':'s','Т':'t',
        'Ћ':'c','У':'u','Ф':'f','Х':'h','Ц':'c','Ч':'c','Џ':'dz','Ш':'s',
    }
    out = "".join(cyr.get(ch, ch) for ch in text)
    out = unicodedata.normalize("NFKD", out)
    out = "".join(c for c in out if not unicodedata.combining(c))
    return out.lower().strip()


def _slugify(text: str) -> str:
    t = _normalize(text)
    t = re.sub(r"[^a-z0-9]+", "-", t).strip("-")
    return t[:80] if t else "party"


def _ensure_dirs():
    os.makedirs(os.path.join(DATA_DIR, "raw", "opendata"), exist_ok=True)


class OpenDataScraper:
    def __init__(self):
        _ensure_dirs()
        self.client = httpx.Client(
            timeout=SCRAPER_TIMEOUT,
            follow_redirects=True,
            headers={"User-Agent": USER_AGENT},
        )

    def _get_dataset(self, slug: str) -> Dict:
        url = f"{UDATA_API}/datasets/{slug}/"
        r = self.client.get(url)
        r.raise_for_status()
        return r.json()

    @staticmethod
    def _pick_newest_resource(resources: List[Dict]) -> Tuple[Dict, str]:
        """
        Choose newest resource based on (created_at/last_modified) when present.
        Prefer xlsx > csv if timestamps are comparable.
        """
        def ts(res: Dict) -> str:
            return (res.get("last_modified") or res.get("created_at") or "").strip()

        def fmt(res: Dict) -> str:
            return (res.get("format") or res.get("mime") or "").lower()

        ranked: List[Tuple[str, int, Dict]] = []
        for res in resources or []:
            f = fmt(res)
            priority = 0
            if "xlsx" in f:
                priority = 2
            elif "csv" in f:
                priority = 1
            ranked.append((ts(res), priority, res))

        ranked.sort(key=lambda x: (x[0], x[1]), reverse=True)
        best = ranked[0][2] if ranked else {}
        best_fmt = fmt(best)
        return best, ("xlsx" if "xlsx" in best_fmt else "csv" if "csv" in best_fmt else best_fmt or "unknown")

    def _download_cached(self, url: str, cache_name: str, force_refresh: bool) -> bytes:
        cache_path = os.path.join(DATA_DIR, "raw", "opendata", cache_name)
        if not force_refresh and os.path.exists(cache_path) and os.path.getsize(cache_path) > 1000:
            logger.info("opendata_cache_hit", file=cache_name)
            return open(cache_path, "rb").read()

        logger.info("opendata_download", url=url, file=cache_name)
        r = self.client.get(url)
        r.raise_for_status()
        data = r.content
        with open(cache_path, "wb") as f:
            f.write(data)
        time.sleep(SCRAPE_DELAY)
        return data

    def fetch_political_parties(self, force_refresh: bool = False) -> List[PoliticalPartyRecord]:
        ds = self._get_dataset(PARTIES_DATASET_SLUG)
        resources = ds.get("resources") or []
        if not resources:
            logger.warning("parties_no_resources", slug=PARTIES_DATASET_SLUG)
            return []

        res, fmt = self._pick_newest_resource(resources)

        res_url = (res.get("url") or "").strip()
        if not res_url:
            logger.warning("parties_resource_missing_url", slug=PARTIES_DATASET_SLUG)
            return []

        # Store metadata for debugging
        meta_path = os.path.join(DATA_DIR, "raw", "opendata", "parties_source.json")
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(
                {
                    "dataset_slug": PARTIES_DATASET_SLUG,
                    "dataset_title": ds.get("title"),
                    "picked_resource": {
                        "title": res.get("title"),
                        "format": res.get("format"),
                        "url": res_url,
                        "created_at": res.get("created_at"),
                        "last_modified": res.get("last_modified"),
                        "filesize": res.get("filesize"),
                    },
                    "scraped_at": datetime.utcnow().isoformat(),
                },
                f,
                ensure_ascii=False,
                indent=2,
            )

        payload = self._download_cached(res_url, f"political_parties.{fmt}", force_refresh=force_refresh)

        records: List[PoliticalPartyRecord] = []
        if fmt == "xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                rowd = dict(zip(headers, row))
                name = str(rowd.get("Странка") or rowd.get("Stranka") or rowd.get("Назив") or rowd.get("Naziv") or "").strip()
                if not name or name == "None":
                    continue

                abbr = str(rowd.get("Скраћени назив") or rowd.get("Skraceni naziv") or rowd.get("Скраћено") or "").strip()
                abbr = abbr if abbr and abbr != "None" else None

                dissolved = str(rowd.get("Датум доношења решења о брисању") or rowd.get("Datum brisanja") or "").strip()
                dissolved = dissolved if dissolved and dissolved != "None" else None
                if dissolved:
                    continue

                pid = f"PARTY-{_slugify(name)}"
                records.append(
                    PoliticalPartyRecord(
                        party_id=pid,
                        name=name,
                        name_normalized=_normalize(name),
                        abbreviation=abbr,
                        founded=str(rowd.get("Датум оснивања") or rowd.get("Datum osnivanja") or "").strip() or None,
                        address=str(rowd.get("Адреса") or rowd.get("Adresa") or "").strip() or None,
                        city=str(rowd.get("Место") or rowd.get("Mesto") or "").strip() or None,
                        leader=str(rowd.get("Заступник") or rowd.get("Zastupnik") or "").strip() or None,
                        dissolved=None,
                    )
                )
            wb.close()

        elif fmt == "csv":
            import csv
            text = payload.decode("utf-8", errors="replace")
            reader = csv.DictReader(text.splitlines())
            for row in reader:
                name = (row.get("Странка") or row.get("Stranka") or row.get("Назив") or row.get("Naziv") or "").strip()
                if not name:
                    continue
                dissolved = (row.get("Датум доношења решења о брисању") or row.get("Datum brisanja") or "").strip() or None
                if dissolved:
                    continue
                pid = f"PARTY-{_slugify(name)}"
                abbr = (row.get("Скраћени назив") or row.get("Skraceni naziv") or "").strip() or None
                records.append(
                    PoliticalPartyRecord(
                        party_id=pid,
                        name=name,
                        name_normalized=_normalize(name),
                        abbreviation=abbr,
                        founded=(row.get("Датум оснивања") or row.get("Datum osnivanja") or "").strip() or None,
                        address=(row.get("Адреса") or row.get("Adresa") or "").strip() or None,
                        city=(row.get("Место") or row.get("Mesto") or "").strip() or None,
                        leader=(row.get("Заступник") or row.get("Zastupnik") or "").strip() or None,
                        dissolved=None,
                    )
                )
        else:
            logger.warning("parties_unknown_format", fmt=fmt)
            return []

        out_path = os.path.join(DATA_DIR, "raw", "opendata", "parties.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump([asdict(r) for r in records], f, ensure_ascii=False, indent=2)

        logger.info("opendata_parties_done", count=len(records), fmt=fmt)
        return records
