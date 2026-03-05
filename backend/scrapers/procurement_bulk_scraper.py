"""
UJN Procurement OpenData Bulk Scraper

Goals:
- Pull newest available years automatically (PROCUREMENT_YEARS=last:2 by default).
- Download OpenData_<YEAR>.xlsx or OpenData_<YEAR>.csv from portal.ujn.gov.rs/OpenD
- Cache downloads, allow force_refresh.
- Parse to JSON outputs expected by GraphLoader:
  - data/raw/institutions/ujn_<YEAR>.json
  - data/raw/ujn/procurements_<YEAR>.json

Important:
- Column names vary. We use normalized header matching and multiple fallbacks.
- If supplier fields exist in the OpenData file, we store:
  supplier_name, supplier_mb
so GraphLoader can create Company and WON_CONTRACT relationships.
"""

import os
import io
import re
import csv
import json
import time
import unicodedata
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any

import httpx
import structlog

logger = structlog.get_logger()

DATA_DIR = os.getenv("DATA_DIR", "./data")
SCRAPE_DELAY = float(os.getenv("SCRAPE_DELAY", "2"))
SCRAPER_TIMEOUT = float(os.getenv("SCRAPER_TIMEOUT", "180"))
USER_AGENT = os.getenv("SCRAPER_UA", "SrpskaTransparentnost/1.0 (+local)")

BASE = os.getenv("UJN_OPEND_BASE", "https://portal.ujn.gov.rs/OpenD").rstrip("/")
BASE_FALLBACK = os.getenv("UJN_OPEND_BASE_FALLBACK", "http://portal.ujn.gov.rs/OpenD").rstrip("/")
PROCUREMENT_YEARS = os.getenv("PROCUREMENT_YEARS", "last:3")
PROCUREMENT_MAX_ROWS = int(os.getenv("PROCUREMENT_MAX_ROWS", "1000"))


def _ensure_dirs():
    os.makedirs(os.path.join(DATA_DIR, "raw", "ujn"), exist_ok=True)
    os.makedirs(os.path.join(DATA_DIR, "raw", "institutions"), exist_ok=True)
    os.makedirs(os.path.join(DATA_DIR, "raw", "cache"), exist_ok=True)


def _norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _digits_only(s: str) -> str:
    if s is None:
        return ""
    return re.sub(r"\D+", "", str(s))


def _resolve_years(spec: str) -> List[str]:
    spec = (spec or "").strip()
    now_year = datetime.utcnow().year
    if spec.startswith("last:"):
        n = int(spec.split(":", 1)[1])
        return [str(now_year - i) for i in range(n)]
    years = [y.strip() for y in spec.split(",") if y.strip()]
    years = sorted(set(years), reverse=True)
    return years


def _candidate_urls(year: str) -> List[Tuple[str, str]]:
    """Return candidate download URLs in priority order (https first, then http fallback)."""
    return [
        (f"{BASE}/OpenData_{year}.xlsx", "xlsx"),
        (f"{BASE}/OpenData_{year}.csv", "csv"),
        (f"{BASE_FALLBACK}/OpenData_{year}.xlsx", "xlsx"),
        (f"{BASE_FALLBACK}/OpenData_{year}.csv", "csv"),
    ]


class ProcurementBulkScraper:
    def __init__(self):
        _ensure_dirs()
        self.client = httpx.Client(
            timeout=SCRAPER_TIMEOUT,
            follow_redirects=True,
            headers={"User-Agent": USER_AGENT},
        )

    def _download_cached(self, url: str, cache_name: str, force_refresh: bool) -> bytes:
        cache_path = os.path.join(DATA_DIR, "raw", "cache", cache_name)
        if not force_refresh and os.path.exists(cache_path) and os.path.getsize(cache_path) > 1000:
            logger.info("ujn_cache_hit", file=cache_name)
            return open(cache_path, "rb").read()

        logger.info("ujn_download", url=url, file=cache_name)
        r = self.client.get(url)
        if r.status_code != 200:
            raise RuntimeError(f"HTTP {r.status_code} for {url}")
        data = r.content
        with open(cache_path, "wb") as f:
            f.write(data)
        time.sleep(SCRAPE_DELAY)
        return data

    def _probe_year(self, year: str) -> Optional[Tuple[str, str, bytes]]:
        for url, fmt in _candidate_urls(year):
            try:
                data = self._download_cached(url, f"OpenData_{year}.{fmt}", force_refresh=False)
                return url, fmt, data
            except Exception:
                continue
        return None

    @staticmethod
    def _find_header(headers_norm: List[str], patterns: List[str]) -> Optional[int]:
        for p in patterns:
            for i, h in enumerate(headers_norm):
                if p in h:
                    return i
        return None

    def _parse_rows(self, fmt: str, payload: bytes, max_rows: int) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Returns: (procurements, institutions)
        """
        rows: List[Dict[str, Any]] = []

        if fmt == "xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
            headers_norm = [_norm(h) for h in headers]

            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                if r_idx > max_rows:
                    break
                rowd = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                rows.append(rowd)

            wb.close()

        elif fmt == "csv":
            text = payload.decode("utf-8", errors="replace")
            reader = csv.DictReader(text.splitlines())
            for r_idx, row in enumerate(reader, start=1):
                if r_idx > max_rows:
                    break
                rows.append(row)

            headers = list(rows[0].keys()) if rows else []
            headers_norm = [_norm(h) for h in headers]
        else:
            raise ValueError(f"Unsupported format: {fmt}")

        # Heuristic column mapping — covers all known UJN OpenData column name variants
        contract_id_idx = self._find_header(headers_norm, [
            "broj obavestenja", "broj obaveštenja",
            "id", "sifra", "šifra", "oznaka",
            "broj postupka", "broj nabavke", "ref. br",
            "redni broj", "rb",
        ])
        title_idx = self._find_header(headers_norm, [
            "predmet nabavke", "predmet", "naziv nabavke", "naziv",
            "opis predmeta", "opis", "subject", "title",
        ])
        inst_name_idx = self._find_header(headers_norm, [
            "naziv narucilaca", "naziv naručioca",
            "narucilac naziv", "naručilac naziv",
            "narucilac", "naručilac",
            "naziv narucilac",
        ])
        inst_mb_idx = self._find_header(headers_norm, [
            "maticni broj narucilaca", "matični broj naručioca",
            "narucilac maticni", "naručilac matični",
            "maticni broj narucioca", "matični broj naručioca",
            "mb narucioca", "mb narucilaca",
        ])
        # Real UJN files use "ugovarac" (contractor) for award records
        supplier_name_idx = self._find_header(headers_norm, [
            "naziv ugovaraca", "naziv ugovarača",
            "ugovarac", "ugovarač",
            "ugovorna strana",
            "ponudjac naziv", "ponuđač naziv",
            "izabrani ponudjac", "izabrani ponuđač",
            "ponudjac", "ponuđač",
            "dobavljac naziv", "dobavljač naziv",
            "dobavljac", "dobavljač",
        ])
        supplier_mb_idx = self._find_header(headers_norm, [
            "maticni broj ugovaraca", "matični broj ugovarača",
            "mb ugovaraca", "mb ugovarača",
            "maticni broj ponudjaca", "matični broj ponuđača",
            "mb ponudjaca", "mb ponuđača",
            "maticni broj dobavljaca", "matični broj dobavljača",
        ])
        value_idx = self._find_header(headers_norm, [
            "ukupna vrednost ugovora", "ukupni iznos ugovora",
            "ukupni iznos", "ukupna vrednost",
            "ugovoreni iznos", "vrednost ugovora",
            "vrednost", "iznos",
        ])
        date_idx = self._find_header(headers_norm, [
            "datum zakljucenja ugovora", "datum zaključenja ugovora",
            "datum zakljucivanja", "datum zaključivanja",
            "datum", "date", "objavljeno",
            "datum obavestenja", "datum obaveštenja",
        ])

        logger.info("ujn_columns_detected",
            contract_id=contract_id_idx,
            title=title_idx,
            inst_name=inst_name_idx,
            supplier_name=supplier_name_idx,
            supplier_mb=supplier_mb_idx,
            value=value_idx,
            date=date_idx,
            total_cols=len(headers),
            sample_cols=headers[:10],
        )

        # Build procurement records
        procurements: List[Dict[str, Any]] = []
        institutions_map: Dict[str, Dict[str, Any]] = {}

        for row in rows:
            # Convert row to list by headers order so idx lookups work for csv too
            row_list = [row.get(h) for h in headers]

            raw_cid = row_list[contract_id_idx] if contract_id_idx is not None else None
            title = row_list[title_idx] if title_idx is not None else None

            # If we cannot identify contract id or title, skip
            cid = str(raw_cid).strip() if raw_cid is not None else ""
            title_s = str(title).strip() if title is not None else ""
            if not cid or not title_s or title_s == "None":
                continue

            inst_name = str(row_list[inst_name_idx]).strip() if inst_name_idx is not None and row_list[inst_name_idx] is not None else ""
            inst_mb = _digits_only(row_list[inst_mb_idx]) if inst_mb_idx is not None else ""

            supplier_name = str(row_list[supplier_name_idx]).strip() if supplier_name_idx is not None and row_list[supplier_name_idx] is not None else ""
            supplier_mb = _digits_only(row_list[supplier_mb_idx]) if supplier_mb_idx is not None else ""

            val_raw = row_list[value_idx] if value_idx is not None else None
            value = None
            if val_raw is not None:
                s = str(val_raw)
                s = s.replace(".", "").replace(",", ".")
                m = re.search(r"([0-9]+(\.[0-9]+)?)", s)
                if m:
                    try:
                        value = float(m.group(1))
                    except Exception:
                        value = None

            date_modified = str(row_list[date_idx]).strip() if date_idx is not None and row_list[date_idx] is not None else ""

            rec = {
                "contract_id": cid,
                "title": title_s,
                "subject": title_s,
                "proc_type": "",
                "subject_type": "",
                "date_modified": date_modified,
                "has_award_decision": True if supplier_name or supplier_mb else False,
                "detail_url": "",
                "institution_mb": inst_mb,
                "institution_name": inst_name,
                "supplier_name": supplier_name or None,
                "supplier_mb": supplier_mb or None,
                "contract_value": value,
                "currency": "RSD",
                "source": "ujn_opend",
            }
            procurements.append(rec)

            if inst_mb and inst_name:
                institutions_map[inst_mb] = {
                    "institution_id": f"UJN-INST-{inst_mb}",
                    "name": inst_name,
                    "name_normalized": _norm(inst_name),
                    "maticni_broj": inst_mb,
                    "pib": "",
                    "source_url": f"http://portal.ujn.gov.rs/Pretrage/Narucilac.aspx?mb={inst_mb}",
                }

        return procurements, list(institutions_map.values())

    def scrape(self, years: Optional[str] = None, max_rows: Optional[int] = None, force_refresh: bool = False) -> Dict[str, Any]:
        years_spec = years or PROCUREMENT_YEARS
        max_rows = max_rows if max_rows is not None else PROCUREMENT_MAX_ROWS

        summary: Dict[str, Any] = {
            "years": [],
            "total_procurements": 0,
            "total_institutions": 0,
            "scraped_at": datetime.utcnow().isoformat(),
        }

        for y in _resolve_years(years_spec):
            found = None
            last_error = None

            for url, fmt in _candidate_urls(y):
                try:
                    payload = self._download_cached(url, f"OpenData_{y}.{fmt}", force_refresh=force_refresh)
                    found = (url, fmt, payload)
                    break
                except Exception as e:
                    last_error = str(e)

            if not found:
                logger.warning("ujn_year_not_available", year=y, error=last_error)
                summary["years"].append({"year": y, "available": False, "error": last_error})
                continue

            url, fmt, payload = found
            procs, insts = self._parse_rows(fmt, payload, max_rows=max_rows)

            out_proc = os.path.join(DATA_DIR, "raw", "ujn", f"procurements_{y}.json")
            out_inst = os.path.join(DATA_DIR, "raw", "institutions", f"ujn_{y}.json")

            with open(out_proc, "w", encoding="utf-8") as f:
                json.dump(procs, f, ensure_ascii=False, indent=2)

            with open(out_inst, "w", encoding="utf-8") as f:
                json.dump(insts, f, ensure_ascii=False, indent=2)

            logger.info("ujn_year_done", year=y, fmt=fmt, procurements=len(procs), institutions=len(insts))
            summary["years"].append(
                {"year": y, "available": True, "fmt": fmt, "url": url, "procurements": len(procs), "institutions": len(insts)}
            )
            summary["total_procurements"] += len(procs)
            summary["total_institutions"] += len(insts)

        return summary
